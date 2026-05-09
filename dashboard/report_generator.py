"""
CascadeGuard Report Generator
Generates a comprehensive multi-sheet Excel report from analysis results.
"""

import io
from datetime import datetime
import networkx as nx
import pandas as pd

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference, PieChart
    from openpyxl.chart.series import DataPoint
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ── Color palette ─────────────────────────────────────────────────────────────
COLORS = {
    "CRITICAL": "DC2626",
    "HIGH":     "D97706",
    "MEDIUM":   "F59E0B",
    "LOW":      "2563EB",
    "CLEAN":    "16A34A",
    "UNKNOWN":  "6B7280",
}

BG = {
    "CRITICAL": "FEE2E2",
    "HIGH":     "FED7AA",
    "MEDIUM":   "FEF3C7",
    "LOW":      "DBEAFE",
    "CLEAN":    "DCFCE7",
    "UNKNOWN":  "F3F4F6",
}

HEADER_DARK  = "1A1F3A"   # navy
HEADER_MID   = "0066CC"   # blue
ACCENT       = "0052A3"
ROW_ALT      = "F8FAFC"
ROW_WHITE    = "FFFFFF"
BORDER_COLOR = "E0E3E8"
TEXT_DARK    = "1A1F3A"
TEXT_GRAY    = "6B7280"


# ── Style helpers ─────────────────────────────────────────────────────────────

def _font(bold=False, size=11, color="1A1F3A", italic=False):
    return Font(name="Arial", bold=bold, size=size,
                color=color, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    side = Side(style="thin", color=BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _set_col_widths(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

def _header_row(ws, row: int, values: list, bg=HEADER_DARK, fg="FFFFFF",
                size=11, height=32):
    ws.row_dimensions[row].height = height
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font      = _font(bold=True, size=size, color=fg)
        cell.fill      = _fill(bg)
        cell.alignment = _center()
        cell.border    = _border()

def _data_row(ws, row: int, values: list, alt=False):
    bg = ROW_ALT if alt else ROW_WHITE
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill      = _fill(bg)
        cell.alignment = _left()
        cell.border    = _border()
        cell.font      = _font(size=10)

def _risk_cell(cell, risk_class: str):
    """Apply risk-colored background + bold text to a cell."""
    rc = risk_class.upper() if risk_class else "UNKNOWN"
    cell.fill  = _fill(BG.get(rc, "F3F4F6"))
    cell.font  = _font(bold=True, size=10, color=COLORS.get(rc, "6B7280"))
    cell.alignment = _center()
    cell.border = _border()


# ═════════════════════════════════════════════════════════════════════════════
# Sheet builders
# ═════════════════════════════════════════════════════════════════════════════

def _build_cover(ws, repo_name: str, df: pd.DataFrame, results: dict):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 22

    # Title banner
    ws.row_dimensions[2].height = 60
    ws.merge_cells("B2:F2")
    title = ws["B2"]
    title.value     = "🛡️  CascadeGuard — Supply Chain Risk Report"
    title.font      = _font(bold=True, size=22, color="FFFFFF")
    title.fill      = _fill(HEADER_DARK)
    title.alignment = _center()

    ws.row_dimensions[3].height = 28
    ws.merge_cells("B3:F3")
    sub = ws["B3"]
    sub.value     = f"Repository: {repo_name}   ·   Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    sub.font      = _font(size=11, color="FFFFFF", italic=True)
    sub.fill      = _fill(HEADER_MID)
    sub.alignment = _center()

    # KPI cards
    total_nodes    = len(df)
    vulnerable     = int((df["cvss_score"] > 0).sum())
    critical_nodes = int((df["risk_class"] == "CRITICAL").sum())
    high_nodes     = int((df["risk_class"] == "HIGH").sum())
    total_vulns    = int(df["vuln_count"].sum())
    fixable        = int(((df["cvss_score"] > 0) & (df["fix_available"])).sum())

    kpis = [
        ("Total Dependencies", total_nodes,    HEADER_DARK),
        ("Vulnerable Packages", vulnerable,    "DC2626"),
        ("Critical Risk",       critical_nodes,"9B1C1C"),
        ("High Risk",           high_nodes,    "B45309"),
        ("Total CVEs",          total_vulns,   ACCENT),
        ("Fixable Now",         fixable,       "16A34A"),
    ]

    cols = ["B", "C", "D", "E", "F"]
    col_map = {0: "B", 1: "C", 2: "D", 3: "E", 4: "F", 5: "F"}  # overflow

    ws.row_dimensions[5].height = 22
    ws.merge_cells("B5:F5")
    ws["B5"].value     = "Executive Summary"
    ws["B5"].font      = _font(bold=True, size=13, color=HEADER_DARK)
    ws["B5"].alignment = _left()

    kpi_row_label = 6
    kpi_row_val   = 7
    ws.row_dimensions[kpi_row_label].height = 22
    ws.row_dimensions[kpi_row_val].height   = 36

    kpi_cols = ["B", "C", "D", "E", "F"]
    for i, (label, val, color) in enumerate(kpis[:5]):
        c = kpi_cols[i]
        lc = ws[f"{c}{kpi_row_label}"]
        lc.value     = label
        lc.font      = _font(size=10, color=TEXT_GRAY)
        lc.alignment = _center()
        lc.fill      = _fill("F8FAFC")
        lc.border    = _border()

        vc = ws[f"{c}{kpi_row_val}"]
        vc.value     = val
        vc.font      = _font(bold=True, size=20, color=color)
        vc.alignment = _center()
        vc.fill      = _fill("F8FAFC")
        vc.border    = _border()

    # Risk distribution table
    ws.row_dimensions[9].height = 22
    ws.merge_cells("B9:F9")
    ws["B9"].value     = "Risk Distribution"
    ws["B9"].font      = _font(bold=True, size=13, color=HEADER_DARK)
    ws["B9"].alignment = _left()

    _header_row(ws, 10,
                ["Risk Class", "Package Count", "% of Total",
                 "Total CVEs", "Avg CVSS"],
                bg=HEADER_DARK)

    rc_order = ["CRITICAL", "HIGH", "MEDIUM", "LOW", "CLEAN"]
    for ridx, rc in enumerate(rc_order):
        rdf = df[df["risk_class"] == rc]
        count = len(rdf)
        pct   = f"{count / total_nodes * 100:.1f}%" if total_nodes else "0%"
        cvss  = f"{rdf['cvss_score'].mean():.1f}" if count else "—"
        cvs_sum = int(rdf["vuln_count"].sum())

        row = 11 + ridx
        ws.row_dimensions[row].height = 22
        _data_row(ws, row,
                  [rc, count, pct, cvs_sum, cvss],
                  alt=(ridx % 2 == 1))
        _risk_cell(ws.cell(row=row, column=1), rc)

    # Methodology note
    ws.row_dimensions[18].height = 18
    ws.merge_cells("B18:F18")
    ws["B18"].value = "Methodology"
    ws["B18"].font  = _font(bold=True, size=12, color=HEADER_DARK)

    methods = [
        ("Dependency Graph", "Full transitive graph up to 3 levels deep across PyPI, npm, Maven"),
        ("Vulnerability DB",  "OSV (Google Open Source Vulnerability) database, queried live"),
        ("Risk Scoring",      "Composite: CVSS × PageRank centrality × depth factor × blast radius"),
        ("Monte Carlo",       "1000 attack propagation simulations per vulnerable node"),
        ("Optimization",      "0/1 Knapsack algorithm to maximize risk reduction within budget"),
    ]
    for midx, (name, desc) in enumerate(methods):
        row = 19 + midx
        ws.row_dimensions[row].height = 20
        nc = ws.cell(row=row, column=2, value=name)
        nc.font = _font(bold=True, size=10, color=HEADER_MID)
        nc.fill = _fill(ROW_ALT if midx % 2 else ROW_WHITE)
        nc.border = _border()
        nc.alignment = _left()

        dc = ws.cell(row=row, column=3, value=desc)
        ws.merge_cells(f"C{row}:F{row}")
        dc.font = _font(size=10, color=TEXT_DARK)
        dc.fill = _fill(ROW_ALT if midx % 2 else ROW_WHITE)
        dc.border = _border()
        dc.alignment = _left()


def _build_all_packages(ws, df: pd.DataFrame):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = [
        "Package", "Version", "Ecosystem", "Depth", "Direct?",
        "Risk Class", "Risk Score", "CVSS Score", "CVE Count",
        "Fix Available", "Dependents (in-degree)", "PageRank",
        "Mean Blast Radius", "Exposure Score", "Explanation"
    ]
    _header_row(ws, 1, headers, bg=HEADER_DARK, height=36)

    _set_col_widths(ws, {
        "A": 32, "B": 14, "C": 12, "D": 8,  "E": 10,
        "F": 12, "G": 12, "H": 12, "I": 10, "J": 14,
        "K": 22, "L": 12, "M": 18, "N": 16, "O": 55,
    })

    for ridx, row in df.iterrows():
        r = ridx + 2 if isinstance(ridx, int) else df.index.get_loc(ridx) + 2
        r = df.index.get_loc(ridx) + 2

        alt = (df.index.get_loc(ridx) % 2 == 1)
        vals = [
            row.get("package", ""),
            row.get("version", "—"),
            row.get("ecosystem", "—"),
            row.get("depth", 0),
            "Yes" if row.get("direct") else "No",
            row.get("risk_class", "CLEAN"),
            round(float(row.get("risk_score", 0)), 2),
            round(float(row.get("cvss_score", 0)), 1),
            int(row.get("vuln_count", 0)),
            "Yes" if row.get("fix_available") else "No",
            int(row.get("in_degree", 0)),
            round(float(row.get("pagerank", 0)), 6),
            round(float(row.get("mean_blast_radius", 0)), 2),
            round(float(row.get("exposure_score", 0)), 2),
            str(row.get("explanation", "")),
        ]
        _data_row(ws, r, vals, alt=alt)
        _risk_cell(ws.cell(row=r, column=6),
                   str(row.get("risk_class", "CLEAN")))

        # Color CVSS cell
        cvss_cell = ws.cell(row=r, column=8)
        cvss = float(row.get("cvss_score", 0))
        if cvss >= 9.0:
            cvss_cell.font = _font(bold=True, size=10, color="DC2626")
        elif cvss >= 7.0:
            cvss_cell.font = _font(bold=True, size=10, color="D97706")
        cvss_cell.alignment = _center()

        # Boolean columns centered
        for col in [5, 10]:
            ws.cell(row=r, column=col).alignment = _center()

        ws.row_dimensions[r].height = 20

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _build_critical_high(ws, df: pd.DataFrame, G):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    sub = df[df["risk_class"].isin(["CRITICAL", "HIGH"])].copy()

    headers = [
        "Package", "Risk Class", "CVSS", "Risk Score",
        "CVE Count", "Fix Available", "Fixed In",
        "Blast Radius (mean)", "Exposure Score",
        "Dependents", "Explanation"
    ]
    _header_row(ws, 1, headers, bg="9B1C1C", fg="FFFFFF", height=36)
    _set_col_widths(ws, {
        "A": 32, "B": 12, "C": 10, "D": 12, "E": 10,
        "F": 14, "G": 14, "H": 18, "I": 16, "J": 12, "K": 55,
    })

    for ridx, (_, row) in enumerate(sub.iterrows()):
        r   = ridx + 2
        alt = ridx % 2 == 1
        node = row.get("package", "")
        ndata = G.nodes.get(node, {}) if G else {}
        vulns = ndata.get("vulnerabilities", [])
        fixed_in = "—"
        if vulns:
            fixes = [v.get("fixed_in") for v in vulns if v.get("fixed_in")]
            fixed_in = fixes[0] if fixes else "No fix"

        vals = [
            node,
            row.get("risk_class", ""),
            round(float(row.get("cvss_score", 0)), 1),
            round(float(row.get("risk_score", 0)), 2),
            int(row.get("vuln_count", 0)),
            "Yes" if row.get("fix_available") else "No",
            fixed_in,
            round(float(row.get("mean_blast_radius", 0)), 2),
            round(float(row.get("exposure_score", 0)), 2),
            int(row.get("in_degree", 0)),
            str(row.get("explanation", "")),
        ]
        _data_row(ws, r, vals, alt=alt)
        _risk_cell(ws.cell(row=r, column=2), str(row.get("risk_class", "")))
        ws.row_dimensions[r].height = 22

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _build_simulation(ws, results: dict, G):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = [
        "Package", "Risk Class", "Exposure Score",
        "Mean Blast Radius", "P95 Blast Radius",
        "Max Blast Radius", "Critical Hit Rate (%)",
        "CVSS Score", "Explanation"
    ]
    _header_row(ws, 1, headers, bg=HEADER_DARK, height=36)
    _set_col_widths(ws, {
        "A": 32, "B": 12, "C": 16, "D": 18, "E": 18,
        "F": 18, "G": 20, "H": 12, "I": 55,
    })

    sim_rows = [
        (node, data) for node, data in results.items()
        if isinstance(data, dict) and "exposure_score" in data
    ]
    sim_rows.sort(key=lambda x: x[1].get("exposure_score", 0), reverse=True)

    for ridx, (node, data) in enumerate(sim_rows):
        r   = ridx + 2
        alt = ridx % 2 == 1
        rc  = G.nodes[node].get("risk_class", "CLEAN") if G and node in G.nodes else "UNKNOWN"

        vals = [
            node,
            rc,
            round(data.get("exposure_score", 0), 2),
            round(data.get("mean_blast_radius", 0), 2),
            data.get("p95_blast_radius", 0),
            data.get("max_blast_radius", 0),
            round(data.get("critical_hit_rate", 0) * 100, 1),
            round(G.nodes[node].get("cvss_score", 0) if G and node in G.nodes else 0, 1),
            G.nodes[node].get("explanation", "") if G and node in G.nodes else "",
        ]
        _data_row(ws, r, vals, alt=alt)
        _risk_cell(ws.cell(row=r, column=2), rc)
        ws.row_dimensions[r].height = 22

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _build_mitigation(ws, plan_data: dict, G):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    all_items = plan_data.get("all_items_ranked", []) if plan_data else []
    selected  = plan_data.get("selected_fixes", []) if plan_data else []
    selected_nodes = {i.get("node") for i in selected}

    headers = [
        "Priority", "Package", "Risk Class", "CVSS Score",
        "Est. Fix Cost (hrs)", "Fix Available", "Fixed In",
        "Risk Score", "Blast Radius", "In Optimization Plan", "Notes"
    ]
    _header_row(ws, 1, headers, bg="065F46", fg="FFFFFF", height=36)
    _set_col_widths(ws, {
        "A": 10, "B": 32, "C": 12, "D": 12, "E": 18,
        "F": 14, "G": 14, "H": 12, "I": 16, "J": 22, "K": 40,
    })

    for ridx, item in enumerate(all_items):
        r   = ridx + 2
        alt = ridx % 2 == 1
        node = item.get("node", "")
        rc   = item.get("risk_class", "")
        in_plan = node in selected_nodes
        fixed_in = item.get("fixed_in") or "No fix available"
        ndata = G.nodes.get(node, {}) if G else {}

        vals = [
            ridx + 1,
            node,
            rc,
            round(float(item.get("cvss_score", 0)), 1),
            round(float(item.get("cost_hours", 0)), 1),
            "Yes" if item.get("fix_available") else "No",
            fixed_in,
            round(float(item.get("risk_score", 0)), 2),
            round(float(ndata.get("mean_blast_radius", 0)), 2),
            "✅ Selected" if in_plan else "—",
            ndata.get("explanation", ""),
        ]
        _data_row(ws, r, vals, alt=alt)
        _risk_cell(ws.cell(row=r, column=3), rc)

        plan_cell = ws.cell(row=r, column=10)
        if in_plan:
            plan_cell.font = _font(bold=True, size=10, color="16A34A")
        ws.row_dimensions[r].height = 22

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _build_cve_detail(ws, df: pd.DataFrame, G):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = [
        "Package", "Ecosystem", "CVE / GHSA ID",
        "Severity", "CVSS Score", "Summary",
        "Fix Version", "Detail URL"
    ]
    _header_row(ws, 1, headers, bg=HEADER_DARK, height=36)
    _set_col_widths(ws, {
        "A": 28, "B": 12, "C": 22, "D": 12, "E": 12,
        "F": 55, "G": 16, "H": 45,
    })

    row = 2
    for _, pkg_row in df.iterrows():
        node  = pkg_row.get("package", "")
        ndata = G.nodes.get(node, {}) if G else {}
        vulns = ndata.get("vulnerabilities", [])
        eco   = pkg_row.get("ecosystem", "—")

        if not vulns:
            continue

        for v in vulns:
            alt = row % 2 == 0
            sev = str(v.get("severity", "UNKNOWN")).upper()
            vals = [
                node,
                eco,
                v.get("display_id", "—"),
                sev,
                round(float(v.get("cvss_score", 0)), 1),
                v.get("summary", "—"),
                v.get("fixed_in") or "No fix",
                v.get("detail_url", ""),
            ]
            _data_row(ws, row, vals, alt=alt)
            _risk_cell(ws.cell(row=row, column=4), sev)
            ws.row_dimensions[row].height = 22
            row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _build_impact(ws, impact_map: dict, G):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    headers = [
        "Package", "Risk Class", "Affected Files",
        "Entry Points", "Features at Risk", "Exposure Scope"
    ]
    _header_row(ws, 1, headers, bg="1E3A5F", fg="FFFFFF", height=36)
    _set_col_widths(ws, {
        "A": 28, "B": 12, "C": 45, "D": 45, "E": 40, "F": 16,
    })

    for ridx, (node_id, idata) in enumerate(impact_map.items()):
        r   = ridx + 2
        alt = ridx % 2 == 1
        rc  = G.nodes[node_id].get("risk_class", "UNKNOWN") if G and node_id in G.nodes else "UNKNOWN"

        files  = ", ".join(idata.get("affected_files", [])[:10])
        eps    = ", ".join(idata.get("entry_points", [])[:8])
        feats  = ", ".join(idata.get("features", []))
        scope  = idata.get("exposure_scope", "MEDIUM")

        vals = [node_id, rc, files, eps, feats, scope]
        _data_row(ws, r, vals, alt=alt)
        _risk_cell(ws.cell(row=r, column=2), rc)
        exp_cell = ws.cell(row=r, column=6)
        scope_color = {"HIGH": "DC2626", "MEDIUM": "D97706", "LOW": "16A34A"}.get(scope, "6B7280")
        exp_cell.font = _font(bold=True, size=10, color=scope_color)
        exp_cell.alignment = _center()
        ws.row_dimensions[r].height = 22

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


# ═════════════════════════════════════════════════════════════════════════════
# Public entry point
# ═════════════════════════════════════════════════════════════════════════════

def generate_report(
    G,
    results: dict,
    df: pd.DataFrame,
    plan_data: dict,
    impact_map: dict,
    repo_name: str,
) -> bytes:
    """
    Generate a comprehensive multi-sheet Excel report.
    Returns raw bytes suitable for st.download_button.
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required: pip install openpyxl")

    wb = Workbook()

    # ── Sheet 1: Cover ────────────────────────────────────────────────────
    ws_cover = wb.active
    ws_cover.title = "📋 Executive Summary"
    ws_cover.sheet_view.showGridLines = False
    _build_cover(ws_cover, repo_name, df, results)

    # ── Sheet 2: All Packages ─────────────────────────────────────────────
    ws_all = wb.create_sheet("📦 All Packages")
    _build_all_packages(ws_all, df)

    # ── Sheet 3: Critical & High ──────────────────────────────────────────
    ws_ch = wb.create_sheet("🚨 Critical & High Risk")
    _build_critical_high(ws_ch, df, G)

    # ── Sheet 4: CVE Detail ───────────────────────────────────────────────
    ws_cve = wb.create_sheet("🔍 CVE Detail")
    _build_cve_detail(ws_cve, df, G)

    # ── Sheet 5: Simulation ───────────────────────────────────────────────
    ws_sim = wb.create_sheet("🎲 Simulation Results")
    _build_simulation(ws_sim, results, G)

    # ── Sheet 6: Mitigation Plan ──────────────────────────────────────────
    ws_mit = wb.create_sheet("⚡ Mitigation Plan")
    _build_mitigation(ws_mit, plan_data, G)

    # ── Sheet 7: Module Impact ────────────────────────────────────────────
    if impact_map:
        ws_imp = wb.create_sheet("💥 Module Impact")
        _build_impact(ws_imp, impact_map, G)

    # ── Save to bytes ─────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()